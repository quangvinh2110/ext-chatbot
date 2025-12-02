from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime
import pandas as pd
import numpy as np
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Set, Tuple, Union
from enum import Enum


class MergeType(Enum):
    """
    Enumeration for merged cell types in Excel.
    
    - NONE: No merge (single cell)
    - HORIZONTAL: Cells merged across columns (same row)
    - VERTICAL: Cells merged across rows (same column)
    - BOTH: Cells merged both horizontally and vertically (rectangular region)
    """
    NONE = "none"
    HORIZONTAL = "horizontal"
    VERTICAL = "vertical"
    BOTH = "both"


class CellType(Enum):
    """
    Enumeration for cell value types used in column type matching.
    """
    NULL = "null"
    STRING = "string"
    NUMBER = "number"
    BOOLEAN = "boolean"
    DATETIME = "datetime"


@dataclass
class MergedCellInfo:
    """
    Data class containing information about a merged cell range in Excel.
    """
    range_str: str
    min_row: int
    max_row: int
    min_col: int
    max_col: int
    value: Any
    rows_spanned: int
    cols_spanned: int
    merge_type: MergeType
    
    def contains_cell(self, row: int, col: int) -> bool:
        """Check if a cell position is within this merged range."""
        return (self.min_row <= row <= self.max_row and 
                self.min_col <= col <= self.max_col)


@dataclass
class RawTable:
    """
    Intermediate representation of a detected table before header/footer identification.
    
    Attributes:
        data: 2D numpy array of table data (without header/footer)
        start_row: Starting row index in the original sheet (0-based)
        end_row: Ending row index in the original sheet (0-based, inclusive)
        start_col: Starting column index (0-based)
        end_col: Ending column index (0-based, inclusive)
        merged_cells: List of merged cell info within this table
        column_types: List of detected types for each column
    """
    data: np.ndarray
    start_row: int
    end_row: int
    start_col: int
    end_col: int
    merged_cells: List[MergedCellInfo] = field(default_factory=list)
    column_types: List[CellType] = field(default_factory=list)
    
    @property
    def num_rows(self) -> int:
        return self.data.shape[0] if self.data.size else 0
    
    @property
    def num_cols(self) -> int:
        return self.end_col - self.start_col + 1
    
    @property
    def shape(self) -> Tuple[int, int]:
        return self.data.shape if self.data.size else (0, 0)


@dataclass
class TableInfo:
    """
    Data class containing a fully parsed table with header and footer identified.
    
    Attributes:
        data: 2D numpy array of table body data (excluding header/footer)
        start_row: Starting row index in the original sheet (0-based)
        end_row: Ending row index in the original sheet (0-based, inclusive)
        start_col: Starting column index (0-based)
        end_col: Ending column index (0-based, inclusive)
        merged_cells: List of merged cell info within this table
        header: Optional 2D numpy array of header rows
        footer: Optional 2D numpy array of footer rows
    """
    data: np.ndarray
    start_row: int
    end_row: int
    start_col: int
    end_col: int
    merged_cells: List[MergedCellInfo] = field(default_factory=list)
    header: Optional[np.ndarray] = None
    footer: Optional[np.ndarray] = None
    
    @property
    def num_rows(self) -> int:
        return self.data.shape[0] if self.data.size else 0
    
    @property
    def num_cols(self) -> int:
        return self.end_col - self.start_col + 1
    
    @property
    def shape(self) -> Tuple[int, int]:
        return self.data.shape if self.data.size else (0, 0)
    
    @property
    def columns(self) -> List[str]:
        """Get column labels from header or generate synthetic ones."""
        if self.header is not None and self.header.size:
            last_header_row = self.header[-1] if self.header.ndim > 1 else self.header
            return [
                str(col) if col is not None and not pd.isna(col) else f"Column {idx + 1}"
                for idx, col in enumerate(last_header_row)
            ]
        return [f"Column {idx + 1}" for idx in range(self.num_cols)]
    
    def to_dataframe(self, include_header: bool = False, include_footer: bool = False) -> pd.DataFrame:
        """
        Convert the table to a pandas DataFrame.
        
        Args:
            include_header: If True, include header rows in the data
            include_footer: If True, include footer rows in the data
        
        Returns:
            DataFrame with table data, optionally including header/footer as rows
        """
        sections = []
        
        if include_header and self.header is not None and self.header.size:
            header_2d = self.header if self.header.ndim > 1 else self.header.reshape(1, -1)
            sections.append(header_2d)
        
        if self.data.size:
            sections.append(self.data)
        
        if include_footer and self.footer is not None and self.footer.size:
            footer_2d = self.footer if self.footer.ndim > 1 else self.footer.reshape(1, -1)
            sections.append(footer_2d)
        
        if not sections:
            return pd.DataFrame(columns=self.columns)
        
        full_data = np.vstack(sections) if len(sections) > 1 else sections[0]
        
        if include_header:
            return pd.DataFrame(full_data)
        else:
            return pd.DataFrame(full_data, columns=self.columns)


class MergeHandler:
    """
    Handles merged cell detection and provides access to cell values with merged cells filled.
    Works directly with openpyxl to preserve cell types.
    """
    
    def __init__(self, workbook_path: str):
        self.workbook_path = workbook_path
        self.workbook: Optional[Workbook] = None
        self.worksheet: Optional[Worksheet] = None
        self._merged_cells_info: List[MergedCellInfo] = []
        self._cell_value_cache: Dict[Tuple[int, int], Any] = {}
        self._max_row: int = 0
        self._max_col: int = 0
        
    def load_sheet(self, sheet_name: Union[str, int] = 0):
        """Load a specific sheet from the workbook."""
        self.workbook = load_workbook(self.workbook_path, data_only=True)
        
        if isinstance(sheet_name, int):
            self.worksheet = self.workbook.worksheets[sheet_name]
        else:
            self.worksheet = self.workbook[sheet_name]
        
        # Get sheet dimensions
        self._max_row = self.worksheet.max_row if self.worksheet.max_row else 0
        self._max_col = self.worksheet.max_column if self.worksheet.max_column else 0
        
        # Extract merged cell information
        self._merged_cells_info = self._extract_merged_cells_info()
        
        # Build cell value cache with merged cells filled
        self._build_cell_cache()
    
    def _extract_merged_cells_info(self) -> List[MergedCellInfo]:
        """Extract information about all merged cells in the current sheet."""
        if self.worksheet is None:
            raise ValueError("Sheet not loaded. Call load_sheet() first.")
        
        merged_cells_list = []
        
        for merged_range in self.worksheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            cell = self.worksheet.cell(min_row, min_col)
            value = cell.value
            
            rows_spanned = max_row - min_row + 1
            cols_spanned = max_col - min_col + 1
            
            if rows_spanned > 1 and cols_spanned > 1:
                merge_type = MergeType.BOTH
            elif rows_spanned > 1:
                merge_type = MergeType.VERTICAL
            elif cols_spanned > 1:
                merge_type = MergeType.HORIZONTAL
            else:
                merge_type = MergeType.NONE
            
            merged_cells_list.append(MergedCellInfo(
                range_str=str(merged_range),
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                value=value,
                rows_spanned=rows_spanned,
                cols_spanned=cols_spanned,
                merge_type=merge_type
            ))
        
        return merged_cells_list
    
    def _build_cell_cache(self):
        """Build a cache of cell values with merged cells filled."""
        if self.worksheet is None:
            return
        
        self._cell_value_cache = {}
        
        # First, read all actual cell values
        for row_idx in range(1, self._max_row + 1):
            for col_idx in range(1, self._max_col + 1):
                cell = self.worksheet.cell(row_idx, col_idx)
                self._cell_value_cache[(row_idx, col_idx)] = cell.value
        
        # Fill merged cells
        for merged_info in self._merged_cells_info:
            value = merged_info.value
            for row in range(merged_info.min_row, merged_info.max_row + 1):
                for col in range(merged_info.min_col, merged_info.max_col + 1):
                    self._cell_value_cache[(row, col)] = value
    
    def get_cell_value(self, row: int, col: int) -> Any:
        """
        Get cell value at (row, col) with merged cells filled.
        
        Args:
            row: 1-based row index (Excel convention)
            col: 1-based column index (Excel convention)
        
        Returns:
            Cell value, or None if cell doesn't exist
        """
        return self._cell_value_cache.get((row, col), None)
    
    def get_merged_cells_info(self) -> List[MergedCellInfo]:
        """Get information about all merged cells."""
        return self._merged_cells_info
    
    def get_max_dimensions(self) -> Tuple[int, int]:
        """Get maximum row and column indices."""
        return (self._max_row, self._max_col)
    
    def close(self):
        """Close the workbook and release resources."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.worksheet = None
            self._cell_value_cache = {}
            self._merged_cells_info = []
            self._max_row = 0
            self._max_col = 0


class TableDetector:
    """
    Detects and extracts individual tables using type-based matching.
    Works directly with openpyxl cell values to preserve types.
    
    Algorithm:
    1. Walk from top-left, expand right until null cell to find first row
    2. Save type for each cell as column types
    3. For each subsequent row, check type matching and handle width differences
    """
    
    def __init__(self, min_cols: int = 1, min_rows: int = 1):
        self.min_cols = min_cols
        self.min_rows = min_rows
    
    @staticmethod
    def get_cell_type(value: Any) -> CellType:
        """Determine the type of a cell value from openpyxl."""
        if value is None:
            return CellType.NULL
        if isinstance(value, bool):
            return CellType.BOOLEAN
        if isinstance(value, (int, float)):
            # Check for NaN
            if isinstance(value, float) and np.isnan(value):
                return CellType.NULL
            return CellType.NUMBER
        if isinstance(value, datetime):
            return CellType.DATETIME
        if isinstance(value, str):
            if value.strip() == "":
                return CellType.NULL
            # Try to detect datetime strings
            try:
                pd.to_datetime(value)
                return CellType.DATETIME
            except (ValueError, TypeError):
                pass
            return CellType.STRING
        return CellType.STRING  # Default fallback
    
    @staticmethod
    def is_null(value: Any) -> bool:
        """Check if a value is null/empty."""
        if value is None:
            return True
        if isinstance(value, float) and np.isnan(value):
            return True
        if isinstance(value, str) and value.strip() == "":
            return True
        return False
    
    @staticmethod
    def types_compatible(type1: CellType, type2: CellType) -> bool:
        """
        Check if two cell types are compatible.
        NULL is compatible with any type (allows sparse data).
        """
        if type1 == CellType.NULL or type2 == CellType.NULL:
            return True
        return type1 == type2
    
    def detect_tables(self, 
                      merge_handler: MergeHandler, 
                      merged_cells: List[MergedCellInfo]) -> List[RawTable]:
        """
        Detect all tables in the sheet.
        
        Args:
            merge_handler: MergeHandler instance with loaded sheet
            merged_cells: List of merged cell information
        
        Returns:
            List of detected RawTable objects
        """
        tables: List[RawTable] = []
        max_row, max_col = merge_handler.get_max_dimensions()
        
        # Track which cells have been assigned to tables (0-based indices)
        used_cells: Set[Tuple[int, int]] = set()
        
        # Iterate through rows (convert from 1-based Excel to 0-based)
        current_row_excel = 1
        while current_row_excel <= max_row:
            # Find the first non-null cell in this row that hasn't been used
            start_col_excel = self._find_first_non_null_col(
                merge_handler, current_row_excel, max_col, used_cells
            )
            
            if start_col_excel is None:
                current_row_excel += 1
                continue
            
            # Try to extract a table starting here
            raw_table = self._extract_table(
                merge_handler, current_row_excel, start_col_excel, 
                max_row, max_col, used_cells, merged_cells
            )
            
            if raw_table is not None and raw_table.num_rows >= self.min_rows and raw_table.num_cols >= self.min_cols:
                tables.append(raw_table)
                # Mark cells as used (0-based)
                for r in range(raw_table.start_row, raw_table.end_row + 1):
                    for c in range(raw_table.start_col, raw_table.end_col + 1):
                        used_cells.add((r, c))
            
            current_row_excel += 1
        
        return tables
    
    def _find_first_non_null_col(self, 
                                  merge_handler: MergeHandler,
                                  row_excel: int,
                                  max_col: int,
                                  used_cells: Set[Tuple[int, int]]) -> Optional[int]:
        """
        Find the first non-null column in a row that hasn't been used.
        
        Returns 1-based Excel column index, or None if not found.
        """
        row = row_excel - 1  # Convert to 0-based for used_cells check
        for col_excel in range(1, max_col + 1):
            col = col_excel - 1  # Convert to 0-based for used_cells check
            if (row, col) in used_cells:
                continue
            value = merge_handler.get_cell_value(row_excel, col_excel)
            if not self.is_null(value):
                return col_excel
        return None
    
    def _extract_table(self,
                       merge_handler: MergeHandler,
                       start_row_excel: int,
                       start_col_excel: int,
                       max_row: int,
                       max_col: int,
                       used_cells: Set[Tuple[int, int]],
                       merged_cells: List[MergedCellInfo]) -> Optional[RawTable]:
        """
        Extract a single table starting from (start_row_excel, start_col_excel).
        
        Uses type-based matching to determine row membership.
        All Excel indices are 1-based.
        """
        # Step 1: Find first row extent and column types
        end_col_excel = start_col_excel
        column_types = []
        
        for col_excel in range(start_col_excel, max_col + 1):
            value = merge_handler.get_cell_value(start_row_excel, col_excel)
            if self.is_null(value):
                break
            column_types.append(self.get_cell_type(value))
            end_col_excel = col_excel
        
        if not column_types:
            return None
        
        # Collect table rows
        first_row_values = []
        for col_excel in range(start_col_excel, end_col_excel + 1):
            first_row_values.append(merge_handler.get_cell_value(start_row_excel, col_excel))
        table_rows = [first_row_values]
        current_end_col_excel = end_col_excel
        
        # Step 2: Process subsequent rows
        row_excel = start_row_excel + 1
        start_col_0 = start_col_excel - 1   # 0-based for used_cells
        
        while row_excel <= max_row:
            # Check if this row position is already used (0-based)
            row = row_excel - 1
            if (row, start_col_0) in used_cells:
                break
            
            # Expand from start_col_excel until null
            row_values = []
            row_types = []
            col_excel = start_col_excel
            
            while col_excel <= max_col:
                value = merge_handler.get_cell_value(row_excel, col_excel)
                if self.is_null(value):
                    break
                row_values.append(value)
                row_types.append(self.get_cell_type(value))
                col_excel += 1
            
            row_end_col_excel = col_excel - 1 if row_values else start_col_excel - 1
            num_row_cells = len(row_values)
            num_table_cols = current_end_col_excel - start_col_excel + 1
            
            # If row is completely empty at start_col, end the table
            if num_row_cells == 0:
                break
            
            # Case analysis based on row width
            if num_row_cells == num_table_cols:
                # Same width: check type matching
                if self._check_types_match(row_types, column_types):
                    table_rows.append(row_values)
                    self._update_column_types(column_types, row_types)
                    row_excel += 1
                else:
                    break
            
            elif num_row_cells > num_table_cols:
                # Wider row: check if existing columns match types
                existing_types = row_types[:num_table_cols]
                if self._check_types_match(existing_types, column_types):
                    # Expand table width
                    new_cols = num_row_cells - num_table_cols
                    new_col_types = row_types[num_table_cols:]
                    
                    # Fill nulls for previous rows in new columns
                    for prev_row in table_rows:
                        prev_row.extend([None] * new_cols)
                    
                    # Add new row
                    table_rows.append(row_values)
                    
                    # Update column types
                    self._update_column_types(column_types, existing_types)
                    column_types.extend(new_col_types)
                    current_end_col_excel = row_end_col_excel
                    row_excel += 1
                else:
                    break
            
            else:  # num_row_cells < num_table_cols
                # Narrower row: expand to match and check non-null types
                extended_values = list(row_values)
                extended_types = list(row_types)
                
                # Fill remaining columns with nulls
                remaining = num_table_cols - num_row_cells
                extended_values.extend([None] * remaining)
                extended_types.extend([CellType.NULL] * remaining)
                
                # Check if non-null cells match corresponding column types
                if self._check_types_match(row_types, column_types[:num_row_cells]):
                    table_rows.append(extended_values)
                    self._update_column_types(column_types, extended_types)
                    row_excel += 1
                else:
                    break
        
        end_row_excel = start_row_excel + len(table_rows) - 1
        
        if len(table_rows) == 0:
            return None
        
        # Convert to numpy array
        data = np.array(table_rows, dtype=object)
        
        # Convert Excel indices (1-based) to 0-based for RawTable
        start_row_0 = start_row_excel - 1
        end_row_0 = end_row_excel - 1
        start_col_0 = start_col_excel - 1
        end_col_0 = current_end_col_excel - 1
        
        # Find merged cells within this table (merged_cells use 1-based Excel indices)
        table_merged_cells = [
            mc for mc in merged_cells
            if (mc.min_row >= start_row_excel and 
                mc.max_row <= end_row_excel and
                mc.min_col >= start_col_excel and
                mc.max_col <= current_end_col_excel)
        ]
        
        return RawTable(
            data=data,
            start_row=start_row_0,
            end_row=end_row_0,
            start_col=start_col_0,
            end_col=end_col_0,
            merged_cells=table_merged_cells,
            column_types=column_types
        )
    
    def _check_types_match(self, row_types: List[CellType], column_types: List[CellType]) -> bool:
        """Check if row types match column types (NULL is compatible with any type)."""
        for rt, ct in zip(row_types, column_types):
            if not self.types_compatible(rt, ct):
                return False
        return True
    
    def _update_column_types(self, column_types: List[CellType], row_types: List[CellType]):
        """Update column types based on new row (non-NULL types take precedence)."""
        for i, rt in enumerate(row_types):
            if i < len(column_types) and column_types[i] == CellType.NULL and rt != CellType.NULL:
                column_types[i] = rt


class HeaderFooterDetector:
    """
    Detects potential headers and footers for raw tables by examining adjacent rows.
    Works directly with openpyxl cell values.
    """
    
    @staticmethod
    def detect(raw_table: RawTable, merge_handler: MergeHandler) -> TableInfo:
        """
        Convert a RawTable to TableInfo by detecting potential header/footer.
        
        Looks at the row immediately above the table data for header,
        and the row immediately below for footer.
        
        Args:
            raw_table: RawTable with 0-based indices
            merge_handler: MergeHandler instance with loaded sheet
        
        Returns:
            TableInfo with detected header/footer
        """
        header = None
        footer = None
        data = raw_table.data
        data_start_row = raw_table.start_row
        data_end_row = raw_table.end_row
        
        # Convert 0-based to 1-based Excel indices
        start_row_excel = raw_table.start_row + 1
        end_row_excel = raw_table.end_row + 1
        start_col_excel = raw_table.start_col + 1
        end_col_excel = raw_table.end_col + 1
        
        # Check for potential header (row above data)
        if raw_table.start_row > 0:
            header_row_excel = start_row_excel - 1
            potential_header = []
            for col_excel in range(start_col_excel, end_col_excel + 1):
                value = merge_handler.get_cell_value(header_row_excel, col_excel)
                potential_header.append(value)
            
            # Check if this row has non-null values and could be a header
            if HeaderFooterDetector._is_potential_header(potential_header, raw_table):
                header = np.array([potential_header], dtype=object)
                data_start_row = raw_table.start_row - 1  # Keep 0-based
        
        # Check for potential footer (row below data)
        max_row, _ = merge_handler.get_max_dimensions()
        if raw_table.end_row < max_row - 1:
            footer_row_excel = end_row_excel + 1
            potential_footer = []
            for col_excel in range(start_col_excel, end_col_excel + 1):
                value = merge_handler.get_cell_value(footer_row_excel, col_excel)
                potential_footer.append(value)
            
            # Check if this row could be a footer
            if HeaderFooterDetector._is_potential_footer(potential_footer, raw_table):
                footer = np.array([potential_footer], dtype=object)
                data_end_row = raw_table.end_row + 1  # Keep 0-based
        
        return TableInfo(
            data=data,
            start_row=data_start_row,
            end_row=data_end_row,
            start_col=raw_table.start_col,
            end_col=raw_table.end_col,
            merged_cells=raw_table.merged_cells,
            header=header,
            footer=footer
        )
    
    @staticmethod
    def _is_potential_header(row_values: List[Any], raw_table: RawTable) -> bool:
        """
        Check if a row could be a header for the table.
        
        Criteria:
        - Has at least one non-null value
        - Values are primarily strings (typical for headers)
        """
        non_null_count = sum(1 for v in row_values if not TableDetector.is_null(v))
        if non_null_count == 0:
            return False
        
        # Check if most values are strings
        string_count = sum(
            1 for v in row_values 
            if not TableDetector.is_null(v) and TableDetector.get_cell_type(v) == CellType.STRING
        )
        
        # Header typically has mostly string values
        if string_count >= non_null_count * 0.5:
            return True
        
        return False
    
    @staticmethod
    def _is_potential_footer(row_values: List[Any], raw_table: RawTable) -> bool:
        """
        Check if a row could be a footer for the table.
        
        Criteria:
        - Has at least one non-null value
        - Could contain summary values, totals, or notes
        """
        non_null_count = sum(1 for v in row_values if not TableDetector.is_null(v))
        if non_null_count == 0:
            return False
        
        # Check for common footer patterns
        for v in row_values:
            if TableDetector.is_null(v):
                continue
            if isinstance(v, str):
                lower_v = v.lower()
                # Common footer keywords
                if any(kw in lower_v for kw in ['total', 'sum', 'average', 'note', 'source']):
                    return True
        
        return False


class ExcelParser:
    """
    Main Excel parser class that orchestrates the entire parsing pipeline.
    
    Usage:
        >>> parser = ExcelParser("data.xlsx")
        >>> tables = parser.parse_sheet("Sheet1")
        >>> for table in tables:
        ...     df = table.to_dataframe()
        >>> parser.close()
        
        # Or with context manager:
        >>> with ExcelParser("data.xlsx") as parser:
        ...     tables = parser.parse_sheet("Sheet1")
    """
    
    def __init__(self, 
                 file_path: str,
                 min_cols: int = 1,
                 min_rows: int = 1,
                 fill_merged_cells: bool = True):
        """
        Initialize the Excel parser.
        
        Args:
            file_path: Path to the Excel workbook file
            min_cols: Minimum columns required for a valid table
            min_rows: Minimum data rows required for a valid table
            fill_merged_cells: Whether to fill merged cells with their values
        """
        self.file_path = file_path
        self.min_cols = min_cols
        self.min_rows = min_rows
        self.fill_merged_cells = fill_merged_cells
        
        self.merge_handler = MergeHandler(file_path)
        self.table_detector = TableDetector(min_cols=min_cols, min_rows=min_rows)
        
        self._current_sheet: Optional[Union[str, int]] = None
        self._merged_cells: List[MergedCellInfo] = []
        self._raw_tables: List[RawTable] = []
        self._tables: List[TableInfo] = []
    
    def parse_sheet(self, sheet_name: Union[str, int] = 0) -> List[TableInfo]:
        """
        Parse a sheet and extract all tables.
        
        Args:
            sheet_name: Sheet name or index (0-based)
        
        Returns:
            List of TableInfo objects for each detected table
        """
        # Load the sheet (this also builds merged cell cache)
        self.merge_handler.load_sheet(sheet_name)
        self._current_sheet = sheet_name
        
        # Get merged cell information
        self._merged_cells = self.merge_handler.get_merged_cells_info()
        
        # Detect raw tables directly from openpyxl
        self._raw_tables = self.table_detector.detect_tables(
            self.merge_handler,
            self._merged_cells
        )
        
        # Detect headers/footers for each raw table
        self._tables = [
            HeaderFooterDetector.detect(raw_table, self.merge_handler)
            for raw_table in self._raw_tables
        ]
        
        return self._tables
    
    def get_table(self, index: int) -> Optional[TableInfo]:
        """Get a specific table by index."""
        if not self._tables:
            raise ValueError("No tables parsed yet. Call parse_sheet() first.")
        
        if 0 <= index < len(self._tables):
            return self._tables[index]
        return None
    
    def get_all_tables(self) -> List[TableInfo]:
        """Get all parsed tables."""
        if self._tables is None:
            raise ValueError("No tables parsed yet. Call parse_sheet() first.")
        return self._tables
    
    def get_raw_tables(self) -> List[RawTable]:
        """Get raw tables (before header/footer detection)."""
        if self._raw_tables is None:
            raise ValueError("No tables parsed yet. Call parse_sheet() first.")
        return self._raw_tables
    
    def get_merged_cells_info(self) -> List[MergedCellInfo]:
        """Get information about all merged cells in the current sheet."""
        if self._merged_cells is None:
            raise ValueError("No sheet parsed yet. Call parse_sheet() first.")
        return self._merged_cells
    
    def summary(self) -> str:
        """Get a summary of parsed tables."""
        if not self._tables:
            return "No tables parsed yet."
        
        lines = [
            "Excel Parser Summary",
            "=" * 50,
            f"File: {self.file_path}",
            f"Sheet: {self._current_sheet}",
            f"Total tables found: {len(self._tables)}",
            f"Total merged cells: {len(self._merged_cells)}",
            "",
            "Tables:"
        ]
        
        for i, table in enumerate(self._tables, 1):
            header_info = f"Header: {table.header.shape[0]} rows" if table.header is not None else "No header"
            footer_info = f"Footer: {table.footer.shape[0]} rows" if table.footer is not None else "No footer"
            lines.append(
                f"  {i}. Rows: {table.start_row}-{table.end_row} | "
                f"Cols: {table.start_col}-{table.end_col} | "
                f"Data shape: {table.shape} | "
                f"{header_info} | {footer_info} | "
                f"Merged: {len(table.merged_cells)}"
            )
        
        return "\n".join(lines)
    
    def visualize_merged_cells(self, max_display: int = 10) -> str:
        """Visualize merged cell information."""
        if not self._merged_cells:
            return "No merged cells found."
        
        lines = [
            "Merged Cells Information",
            "=" * 50,
            f"Total: {len(self._merged_cells)} merged ranges",
            ""
        ]
        
        for i, merge in enumerate(self._merged_cells[:max_display], 1):
            lines.extend([
                f"{i}. Range: {merge.range_str}",
                f"   Type: {merge.merge_type.value}",
                f"   Rows: {merge.min_row}-{merge.max_row} (spanning {merge.rows_spanned})",
                f"   Cols: {merge.min_col}-{merge.max_col} (spanning {merge.cols_spanned})",
                f"   Value: {merge.value}",
                ""
            ])
        
        if len(self._merged_cells) > max_display:
            lines.append(f"... and {len(self._merged_cells) - max_display} more")
        
        return "\n".join(lines)
    
    def export_table(self, 
                     table_index: int, 
                     output_path: str, 
                     file_format: str = 'csv',
                     include_header: bool = False,
                     include_footer: bool = False):
        """
        Export a specific table to a file.
        
        Args:
            table_index: Zero-based index of the table to export
            output_path: Path for the output file
            file_format: 'csv', 'excel', or 'json'
            include_header: Include header rows in export
            include_footer: Include footer rows in export
        """
        table = self.get_table(table_index)
        if table is None:
            raise ValueError(f"Table index {table_index} not found.")
        
        df = table.to_dataframe(include_header=include_header, include_footer=include_footer)
        
        if file_format == 'csv':
            df.to_csv(output_path, index=False)
        elif file_format == 'excel':
            df.to_excel(output_path, index=False)
        elif file_format == 'json':
            df.to_json(output_path, orient='records', indent=2)
        else:
            raise ValueError(f"Unsupported format: {file_format}")
    
    def close(self):
        """Close the parser and release resources."""
        self.merge_handler.close()
        self._current_sheet = None
        self._merged_cells = []
        self._raw_tables = []
        self._tables = []
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
    
    def __repr__(self):
        if not self._tables:
            return f"ExcelParser(file='{self.file_path}', status='not parsed')"
        return f"ExcelParser(file='{self.file_path}', tables={len(self._tables)})"
