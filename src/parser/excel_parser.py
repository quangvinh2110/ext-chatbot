from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
import numpy as np
from dataclasses import dataclass
from typing import Any, List, Optional, Union, Tuple
from enum import Enum


class MergeType(Enum):
    """
    Enumeration for merged cell types in Excel.
    
    This enum represents the different ways cells can be merged:
    - NONE: No merge (single cell)
    - HORIZONTAL: Cells merged across columns (same row)
    - VERTICAL: Cells merged across rows (same column)
    - BOTH: Cells merged both horizontally and vertically (rectangular region)
    
    Examples:
        >>> merge_type = MergeType.HORIZONTAL
        >>> print(merge_type.value)
        'horizontal'
    """
    NONE = "none"
    HORIZONTAL = "horizontal"
    VERTICAL = "vertical"
    BOTH = "both"


@dataclass
class MergedCellInfo:
    """
    Data class containing information about a merged cell range in Excel.
    
    This class stores all relevant metadata about merged cells including
    their position, dimensions, value, and merge type.
    
    Attributes:
        range_str (str): String representation of the merged range (e.g., "A1:B2").
        min_row (int): Minimum row index (1-based, Excel convention).
        max_row (int): Maximum row index (1-based, Excel convention).
        min_col (int): Minimum column index (1-based, Excel convention).
        max_col (int): Maximum column index (1-based, Excel convention).
        value (any): The value stored in the merged cell (from top-left cell).
        rows_spanned (int): Number of rows spanned by the merge.
        cols_spanned (int): Number of columns spanned by the merge.
        merge_type (MergeType): Type of merge (horizontal, vertical, both, or none).
    
    Methods:
        contains_cell(row, col): Check if a cell position is within this merged range.
    
    Examples:
        >>> merge_info = MergedCellInfo(
        ...     range_str="A1:B2",
        ...     min_row=1, max_row=2,
        ...     min_col=1, max_col=2,
        ...     value="Header",
        ...     rows_spanned=2,
        ...     cols_spanned=2,
        ...     merge_type=MergeType.BOTH
        ... )
        >>> merge_info.contains_cell(1, 1)
        True
        >>> merge_info.contains_cell(3, 1)
        False
    """
    range_str: str
    min_row: int
    max_row: int
    min_col: int
    max_col: int
    value: Any
    rows_spanned: int
    cols_spanned: int
    merge_type: MergeType
    
    def contains_cell(self, row: int, col: int) -> bool:
        """
        Check if a cell position is within this merged range.
        
        Args:
            row (int): Row index (1-based).
            col (int): Column index (1-based).
        
        Returns:
            bool: True if the cell is within the merged range, False otherwise.
        
        Examples:
            >>> merge_info = MergedCellInfo(...)
            >>> merge_info.contains_cell(1, 1)
            True
        """
        return (self.min_row <= row <= self.max_row and 
                self.min_col <= col <= self.max_col)


@dataclass
class TableInfo:
    """
    Data class containing information about an extracted table from Excel.
    
    This class stores the table data along with metadata about its position
    in the original sheet, dimensions, and associated merged cells.
    
    Attributes:
        data (pd.DataFrame): The extracted table as a pandas DataFrame.
        start_row (int): Starting row index in the original sheet (0-based).
        end_row (int): Ending row index in the original sheet (0-based).
        start_col (int): Starting column index in the original sheet (0-based).
        num_cols (int): Number of columns in the table.
        num_data_rows (int): Number of data rows (excluding header).
        merged_cells (List[MergedCellInfo]): List of merged cells within this table.
    
    Properties:
        shape (Tuple[int, int]): Tuple of (rows, columns) dimensions.
    
    Examples:
        >>> table_info = TableInfo(...)
        >>> print(f"Table shape: {table_info.shape}")
        >>> print(table_info.data)
    """
    data: np.ndarray
    start_row: int
    end_row: int
    start_col: int
    num_cols: int
    num_data_rows: int
    merged_cells: List[MergedCellInfo]
    
    @property
    def shape(self) -> Tuple[int, int]:
        """
        Get the shape of the table numpy array.
        
        Returns:
            Tuple[int, int]: A tuple containing (number of rows, number of columns).
        
        Examples:
            >>> table_info.shape
            (10, 5)
        """
        return self.data.shape if self.data.size else (0, 0)


class MergeHandler:
    """
    Handles merged cell detection and filling in Excel files.
    
    This class is responsible for:
    - Loading Excel workbooks and sheets
    - Detecting merged cell ranges
    - Reading sheet data with proper handling of merged cells
    - Filling merged cells with their values
    
    The class uses openpyxl to access Excel file structure and can handle
    merged cells that span horizontally, vertically, or both directions.
    
    Attributes:
        workbook_path (str): Path to the Excel workbook file.
        workbook (openpyxl.workbook.Workbook): The loaded workbook object.
        worksheet (openpyxl.worksheet.worksheet.Worksheet): The current worksheet.
    
    Examples:
        >>> handler = MergeHandler("data.xlsx")
        >>> handler.load_sheet("Sheet1")
        >>> merged_cells = handler.get_merged_cells_info()
        >>> df = handler.read_sheet_with_merged_cells(fill_merged=True)
        >>> handler.close()
    """
    
    def __init__(self, workbook_path: str):
        """
        Initialize the MergeHandler with a workbook path.
        
        Args:
            workbook_path (str): Path to the Excel workbook file (.xlsx format).
        
        Raises:
            FileNotFoundError: If the workbook file does not exist.
        
        Examples:
            >>> handler = MergeHandler("/path/to/file.xlsx")
        """
        self.workbook_path = workbook_path
        self.workbook: Workbook = None
        self.worksheet: Worksheet = None
        
    def load_sheet(self, sheet_name: Union[str, int] = 0):
        """
        Load a specific sheet from the workbook.
        
        This method opens the workbook and selects the specified sheet.
        Must be called before using other methods that require sheet access.
        
        Args:
            sheet_name (Union[str, int], optional): 
                Sheet name as string or index as integer. Defaults to 0 (first sheet).
        
        Raises:
            ValueError: If the sheet name/index is invalid.
            FileNotFoundError: If the workbook file cannot be found.
        
        Examples:
            >>> handler.load_sheet(0)  # Load first sheet by index
            >>> handler.load_sheet("Sheet1")  # Load sheet by name
        """
        self.workbook = load_workbook(self.workbook_path)
        
        if isinstance(sheet_name, int):
            self.worksheet = self.workbook.worksheets[sheet_name]
        else:
            self.worksheet = self.workbook[sheet_name]
    
    def get_merged_cells_info(self) -> List[MergedCellInfo]:
        """
        Extract information about all merged cells in the current sheet.
        
        This method scans the worksheet for all merged cell ranges and returns
        detailed information about each merge, including position, dimensions,
        value, and merge type.
        
        Returns:
            List[MergedCellInfo]: List of MergedCellInfo objects, one for each
                merged cell range found in the sheet.
        
        Raises:
            ValueError: If no sheet has been loaded (call load_sheet() first).
        
        Examples:
            >>> handler.load_sheet("Sheet1")
            >>> merged_cells = handler.get_merged_cells_info()
            >>> print(f"Found {len(merged_cells)} merged ranges")
            >>> for merge in merged_cells:
            ...     print(f"{merge.range_str}: {merge.merge_type.value}")
        """
        if self.worksheet is None:
            raise ValueError("Sheet not loaded. Call load_sheet() first.")
        
        merged_cells_list = []
        
        for merged_range in self.worksheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            value = self.worksheet.cell(min_row, min_col).value
            
            rows_spanned = max_row - min_row + 1
            cols_spanned = max_col - min_col + 1
            
            # Determine merge type
            if rows_spanned > 1 and cols_spanned > 1:
                merge_type = MergeType.BOTH
            elif rows_spanned > 1:
                merge_type = MergeType.VERTICAL
            elif cols_spanned > 1:
                merge_type = MergeType.HORIZONTAL
            else:
                merge_type = MergeType.NONE
            
            merged_cells_list.append(MergedCellInfo(
                range_str=str(merged_range),
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                value=value,
                rows_spanned=rows_spanned,
                cols_spanned=cols_spanned,
                merge_type=merge_type
            ))
        
        return merged_cells_list
    
    def read_sheet_with_merged_cells(self, fill_merged: bool = True) -> pd.DataFrame:
        """
        Read sheet data with merged cells properly handled.
        
        This method reads the entire sheet into a pandas DataFrame. If
        fill_merged is True, all cells within a merged range will be filled
        with the value from the top-left cell. Otherwise, only the top-left
        cell contains the value, and other cells in the merge are None.
        
        Args:
            fill_merged (bool, optional): 
                If True, fill all cells in merged ranges with the merged value.
                If False, only top-left cell has value. Defaults to True.
        
        Returns:
            pd.DataFrame: DataFrame containing the sheet data with merged cells
                handled according to fill_merged parameter.
        
        Raises:
            ValueError: If no sheet has been loaded (call load_sheet() first).
        
        Examples:
            >>> handler.load_sheet("Sheet1")
            >>> # Fill all merged cells with their values
            >>> df = handler.read_sheet_with_merged_cells(fill_merged=True)
            >>> # Keep merged cells as-is (only top-left has value)
            >>> df = handler.read_sheet_with_merged_cells(fill_merged=False)
        """
        if self.worksheet is None:
            raise ValueError("Sheet not loaded. Call load_sheet() first.")
        
        # Get merged cell ranges
        merged_ranges = list(self.worksheet.merged_cells.ranges)
        
        # Read data into list
        data = []
        for row in self.worksheet.iter_rows(values_only=False):
            row_data = [cell.value for cell in row]
            data.append(row_data)
        
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        if fill_merged:
            # Fill merged cells with value from top-left cell
            for merged_range in merged_ranges:
                min_col, min_row, max_col, max_row = merged_range.bounds
                
                # Get value from top-left cell (adjust for 0-based indexing)
                value = df.iloc[min_row - 1, min_col - 1]
                
                # Fill all cells in the merged range
                for row in range(min_row - 1, max_row):
                    for col in range(min_col - 1, max_col):
                        df.iloc[row, col] = value
        
        return df
    
    def close(self):
        """
        Close the workbook and release resources.
        
        This method should be called when done with the handler to properly
        close the workbook file and free system resources. The handler can
        be reused by calling load_sheet() again after closing.
        
        Examples:
            >>> handler = MergeHandler("data.xlsx")
            >>> handler.load_sheet("Sheet1")
            >>> # ... use handler ...
            >>> handler.close()
        """
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.worksheet = None


class TableDetector:
    """
    Detects and extracts individual tables from a DataFrame.
    
    This class implements the algorithm to identify and extract multiple
    tables from a single DataFrame. It uses the following strategy:
    1. Start at the first cell
    2. Find all columns of the current table (stop at first NA in header row)
    3. Find all rows of the current table (stop when row contains all nulls)
    4. Repeat for remaining tables
    
    The detector can be configured with minimum size requirements and
    options for handling empty rows between tables.
    
    Attributes:
        min_cols (int): Minimum number of columns required for a valid table.
        min_rows (int): Minimum number of data rows (excluding header) required.
        skip_empty_rows (bool): Whether to skip completely empty rows between tables.
    
    Examples:
        >>> detector = TableDetector(min_cols=2, min_rows=1)
        >>> tables = detector.detect_tables(df_full, merged_cells)
        >>> print(f"Found {len(tables)} tables")
    """
    
    def __init__(self, 
                 min_cols: int = 1, 
                 min_rows: int = 1, 
                 skip_empty_rows: bool = True):
        """
        Initialize the TableDetector with configuration parameters.
        
        Args:
            min_cols (int, optional): 
                Minimum number of columns required for a table to be considered valid.
                Tables with fewer columns will be skipped. Defaults to 1.
            min_rows (int, optional): 
                Minimum number of data rows (excluding header) required for a table
                to be considered valid. Defaults to 1.
            skip_empty_rows (bool, optional): 
                If True, completely empty rows between tables will be skipped.
                If False, empty rows may be considered as potential table boundaries.
                Defaults to True.
        
        Examples:
            >>> # Strict detector: require at least 3 columns and 2 data rows
            >>> detector = TableDetector(min_cols=3, min_rows=2)
            >>> # Lenient detector: accept any table with at least 1 column
            >>> detector = TableDetector(min_cols=1, min_rows=0)
        """
        self.min_cols = min_cols
        self.min_rows = min_rows
        self.skip_empty_rows = skip_empty_rows
    
    def detect_tables(self, 
                     df_full: pd.DataFrame, 
                     merged_cells: List[MergedCellInfo]) -> List[TableInfo]:
        """
        Detect all tables in the DataFrame.
        
        This method scans the entire DataFrame from top to bottom, identifying
        and extracting individual tables. Each table is detected by:
        1. Finding a header row (first non-empty row)
        2. Determining column boundaries (stop at first NA in header)
        3. Determining row boundaries (stop at first all-NA row)
        4. Validating against minimum size requirements
        
        Args:
            df_full (pd.DataFrame): 
                The full DataFrame containing one or more tables.
            merged_cells (List[MergedCellInfo]): 
                List of merged cell information for associating merges with tables.
        
        Returns:
            List[TableInfo]: List of TableInfo objects, one for each detected table.
                Tables are returned in order of appearance (top to bottom).
        
        Examples:
            >>> detector = TableDetector()
            >>> tables = detector.detect_tables(df_full, merged_cells)
            >>> for i, table in enumerate(tables):
            ...     print(f"Table {i+1}: {table.shape} at rows {table.start_row}-{table.end_row}")
        """
        tables = []
        current_row = 0
        
        while current_row < len(df_full):
            # Skip empty rows if configured
            if self.skip_empty_rows and df_full.iloc[current_row].isna().all():
                current_row += 1
                continue
            
            # Try to extract a table
            table_info = self._extract_table(df_full, current_row, merged_cells)
            
            if table_info is not None:
                tables.append(table_info)
                current_row = table_info.end_row + 1
            else:
                current_row += 1
        
        return tables
    
    def _extract_table(self, 
                      df_full: pd.DataFrame, 
                      start_row: int,
                      merged_cells: List[MergedCellInfo]) -> Optional[TableInfo]:
        """
        Extract a single table starting from start_row.
        
        This is an internal method that implements the core table extraction
        algorithm. It finds the column and row boundaries of a table starting
        at the given row, validates it meets minimum requirements, and returns
        a TableInfo object if successful.
        
        Args:
            df_full (pd.DataFrame): The full DataFrame to extract from.
            start_row (int): The starting row index (0-based) to begin extraction.
            merged_cells (List[MergedCellInfo]): List of merged cell information.
        
        Returns:
            Optional[TableInfo]: TableInfo object if a valid table is found,
                None otherwise. Returns None if:
                - No valid header row found
                - Table doesn't meet minimum column/row requirements
                - Table has no data rows
        
        Examples:
            >>> table_info = detector._extract_table(df_full, 0, merged_cells)
            >>> if table_info:
            ...     print(f"Extracted table with {table_info.num_cols} columns")
        """
        # Step 1: Find columns - read header row until we hit NA
        header_row = df_full.iloc[start_row]
        
        num_cols = 0
        for i, val in enumerate(header_row):
            if pd.isna(val):
                break
            num_cols += 1
        
        # Validate minimum columns
        if num_cols < self.min_cols:
            return None
        
        # Step 2: Find rows - continue until we hit a row with all NAs
        end_row = start_row + 1
        while end_row < len(df_full):
            row_data = df_full.iloc[end_row, :num_cols]
            full_row = df_full.iloc[end_row]

            if row_data.isna().all():
                break

            if self._has_longer_consecutive_values(full_row, num_cols):
                break
            end_row += 1
        
        num_data_rows = end_row - start_row - 1
        
        # Validate minimum rows
        if num_data_rows < self.min_rows:
            return None
        
        # Step 3: Extract the table data
        table_data = df_full.iloc[start_row:end_row, :num_cols]
        
        # Convert to numpy array (includes header row as first row)
        table_array = table_data.values
        
        # Step 4: Find merged cells that belong to this table
        table_merged_cells = [
            mc for mc in merged_cells
            if (mc.min_row - 1 >= start_row and 
                mc.max_row - 1 <= end_row and
                mc.min_col - 1 < num_cols)
        ]
        
        return TableInfo(
            data=table_array,
            start_row=start_row,
            end_row=end_row - 1,
            start_col=0,
            num_cols=num_cols,
            num_data_rows=num_data_rows,
            merged_cells=table_merged_cells
        )

    @staticmethod
    def _has_longer_consecutive_values(row: pd.Series, max_allowed: int) -> bool:
        """
        Check if the row has a run of consecutive non-null values longer than max_allowed.
        This indicates the row is likely the header of the next table.
        """
        consecutive = 0
        for value in row:
            if pd.isna(value):
                consecutive = 0
                continue
            consecutive += 1
            if consecutive > max_allowed:
                return True
        return False


class ExcelParser:
    """
    Main Excel parser class that orchestrates the entire parsing pipeline.
    
    This class provides a high-level interface for parsing Excel files with
    multiple tables and merged cells. It combines MergeHandler and TableDetector
    to provide a complete solution for extracting structured data from Excel sheets.
    
    The parser handles:
    - Loading Excel workbooks and sheets
    - Detecting and handling merged cells (horizontal, vertical, or both)
    - Identifying multiple tables within a single sheet
    - Extracting tables with metadata (position, dimensions, merged cells)
    - Exporting tables to various formats
    
    Attributes:
        file_path (str): Path to the Excel file.
        min_cols (int): Minimum columns required for a valid table.
        min_rows (int): Minimum data rows required for a valid table.
        skip_empty_rows (bool): Whether to skip empty rows between tables.
        fill_merged_cells (bool): Whether to fill merged cells with their values.
        merge_handler (MergeHandler): Handler for merged cell operations.
        table_detector (TableDetector): Detector for table extraction.
    
    Examples:
        Basic usage:
        >>> parser = ExcelParser("data.xlsx")
        >>> tables = parser.parse_sheet("Sheet1")
        >>> print(parser.summary())
        >>> parser.close()
        
        Using context manager:
        >>> with ExcelParser("data.xlsx") as parser:
        ...     tables = parser.parse_sheet("Sheet1")
        ...     for table in tables:
        ...         print(table.data)
        
        Custom configuration:
        >>> parser = ExcelParser(
        ...     "data.xlsx",
        ...     min_cols=3,
        ...     min_rows=2,
        ...     fill_merged_cells=True
        ... )
    """
    
    def __init__(self, 
                 file_path: str,
                 min_cols: int = 1,
                 min_rows: int = 1,
                 skip_empty_rows: bool = True,
                 fill_merged_cells: bool = True):
        """
        Initialize the Excel parser with configuration parameters.
        
        Args:
            file_path (str): 
                Path to the Excel workbook file (.xlsx format).
            min_cols (int, optional): 
                Minimum number of columns required for a table to be considered valid.
                Tables with fewer columns will be skipped. Defaults to 1.
            min_rows (int, optional): 
                Minimum number of data rows (excluding header) required for a table
                to be considered valid. Defaults to 1.
            skip_empty_rows (bool, optional): 
                If True, completely empty rows between tables will be skipped.
                Defaults to True.
            fill_merged_cells (bool, optional): 
                If True, all cells within merged ranges will be filled with the
                merged value. If False, only the top-left cell contains the value.
                Defaults to True.
        
        Raises:
            FileNotFoundError: If the Excel file does not exist.
        
        Examples:
            >>> parser = ExcelParser("data.xlsx")
            >>> parser = ExcelParser("data.xlsx", min_cols=3, min_rows=2)
        """
        self.file_path = file_path
        self.min_cols = min_cols
        self.min_rows = min_rows
        self.skip_empty_rows = skip_empty_rows
        self.fill_merged_cells = fill_merged_cells
        
        # Initialize sub-components
        self.merge_handler = MergeHandler(file_path)
        self.table_detector = TableDetector(
            min_cols=min_cols,
            min_rows=min_rows,
            skip_empty_rows=skip_empty_rows
        )
        
        self._current_sheet: Union[str, int] = 0
        self._df_full: pd.DataFrame = pd.DataFrame()
        self._merged_cells: List[MergedCellInfo] = []
        self._tables: List[TableInfo] = []
    
    def parse_sheet(self, sheet_name: Union[str, int] = 0) -> List[TableInfo]:
        """
        Parse a sheet and extract all tables.
        
        This is the main method that orchestrates the parsing pipeline:
        1. Loads the specified sheet
        2. Detects all merged cells
        3. Reads sheet data with merged cells handled
        4. Detects and extracts all tables
        5. Associates merged cells with their respective tables
        
        Args:
            sheet_name (Union[str, int], optional): 
                Sheet name as string or index as integer. Defaults to 0 (first sheet).
        
        Returns:
            List[TableInfo]: List of TableInfo objects, one for each detected table.
                Tables are returned in order of appearance (top to bottom).
        
        Raises:
            ValueError: If the sheet name/index is invalid.
            FileNotFoundError: If the workbook file cannot be found.
        
        Examples:
            >>> parser = ExcelParser("data.xlsx")
            >>> tables = parser.parse_sheet("Sheet1")
            >>> tables = parser.parse_sheet(0)  # First sheet by index
            >>> print(f"Found {len(tables)} tables")
        """
        # Load the sheet
        self.merge_handler.load_sheet(sheet_name)
        self._current_sheet = sheet_name
        
        # Get merged cell information
        self._merged_cells = self.merge_handler.get_merged_cells_info()
        
        # Read the full sheet with merged cells handled
        self._df_full = self.merge_handler.read_sheet_with_merged_cells(
            fill_merged=self.fill_merged_cells
        )
        
        # Detect and extract tables
        self._tables = self.table_detector.detect_tables(
            self._df_full, 
            self._merged_cells
        )
        
        return self._tables
    
    def get_table(self, index: int) -> Optional[TableInfo]:
        """
        Get a specific table by index.
        
        Args:
            index (int): Zero-based index of the table to retrieve.
        
        Returns:
            Optional[TableInfo]: TableInfo object if index is valid, None otherwise.
        
        Raises:
            ValueError: If parse_sheet() has not been called yet.
        
        Examples:
            >>> parser.parse_sheet("Sheet1")
            >>> first_table = parser.get_table(0)
            >>> if first_table:
            ...     print(first_table.data)
        """
        if self._tables is None:
            raise ValueError("No tables parsed yet. Call parse_sheet() first.")
        
        if 0 <= index < len(self._tables):
            return self._tables[index]
        return None
    
    def get_all_tables(self) -> List[TableInfo]:
        """
        Get all parsed tables.
        
        Returns:
            List[TableInfo]: List of all TableInfo objects from the last parse_sheet() call.
        
        Raises:
            ValueError: If parse_sheet() has not been called yet.
        
        Examples:
            >>> parser.parse_sheet("Sheet1")
            >>> all_tables = parser.get_all_tables()
            >>> print(f"Total tables: {len(all_tables)}")
        """
        if self._tables is None:
            raise ValueError("No tables parsed yet. Call parse_sheet() first.")
        return self._tables
    
    def get_merged_cells_info(self) -> List[MergedCellInfo]:
        """
        Get information about all merged cells in the current sheet.
        
        Returns:
            List[MergedCellInfo]: List of MergedCellInfo objects for all merged
                cells in the last parsed sheet.
        
        Raises:
            ValueError: If parse_sheet() has not been called yet.
        
        Examples:
            >>> parser.parse_sheet("Sheet1")
            >>> merged_cells = parser.get_merged_cells_info()
            >>> print(f"Found {len(merged_cells)} merged cell ranges")
        """
        if self._merged_cells is None:
            raise ValueError("No sheet parsed yet. Call parse_sheet() first.")
        return self._merged_cells
    
    def summary(self) -> str:
        """
        Get a summary of parsed tables.
        
        Returns a formatted string containing:
        - File path and sheet name
        - Total number of tables found
        - Total number of merged cells
        - Details about each table (position, dimensions, merged cells)
        
        Returns:
            str: Formatted summary string.
        
        Raises:
            ValueError: If parse_sheet() has not been called yet.
        
        Examples:
            >>> parser.parse_sheet("Sheet1")
            >>> print(parser.summary())
        """
        if self._tables is None:
            return "No tables parsed yet."
        
        summary_lines = [
            "Excel Parser Summary",
            "=" * 50,
            f"File: {self.file_path}",
            f"Sheet: {self._current_sheet}",
            f"Total tables found: {len(self._tables)}",
            f"Total merged cells: {len(self._merged_cells)}",
            "",
            "Tables:"
        ]
        
        for i, table in enumerate(self._tables, 1):
            summary_lines.append(
                f"  {i}. Rows: {table.start_row}-{table.end_row} | "
                f"Shape: {table.shape} | "
                f"Merged cells: {len(table.merged_cells)}"
            )
        
        return "\n".join(summary_lines)
    
    def visualize_merged_cells(self, max_display: int = 10) -> str:
        """
        Visualize merged cell information.
        
        Returns a formatted string showing details about merged cells including
        their ranges, types, dimensions, and values.
        
        Args:
            max_display (int, optional): 
                Maximum number of merged cells to display. Defaults to 10.
                If there are more, a summary message is shown.
        
        Returns:
            str: Formatted string with merged cell information.
        
        Raises:
            ValueError: If parse_sheet() has not been called yet.
        
        Examples:
            >>> parser.parse_sheet("Sheet1")
            >>> print(parser.visualize_merged_cells())
            >>> print(parser.visualize_merged_cells(max_display=20))
        """
        if self._merged_cells is None:
            return "No merged cells information available."
        
        lines = [
            "Merged Cells Information",
            "=" * 50,
            f"Total: {len(self._merged_cells)} merged ranges",
            ""
        ]
        
        for i, merge in enumerate(self._merged_cells[:max_display], 1):
            lines.extend([
                f"{i}. Range: {merge.range_str}",
                f"   Type: {merge.merge_type.value}",
                f"   Rows: {merge.min_row}-{merge.max_row} (spanning {merge.rows_spanned})",
                f"   Cols: {merge.min_col}-{merge.max_col} (spanning {merge.cols_spanned})",
                f"   Value: {merge.value}",
                ""
            ])
        
        if len(self._merged_cells) > max_display:
            lines.append(f"... and {len(self._merged_cells) - max_display} more")
        
        return "\n".join(lines)
    
    def close(self):
        """
        Close the parser and release resources.
        
        This method closes the workbook file and cleans up resources.
        Should be called when done with the parser. The parser can be
        reused by calling parse_sheet() again after closing.
        
        Examples:
            >>> parser = ExcelParser("data.xlsx")
            >>> parser.parse_sheet("Sheet1")
            >>> # ... use parser ...
            >>> parser.close()
        """
        self.merge_handler.close()
        self._current_sheet = None
        self._df_full = None
        self._merged_cells = None
        self._tables = None
    
    def __enter__(self):
        """
        Context manager entry point.
        
        Returns:
            ExcelParser: The parser instance itself.
        
        Examples:
            >>> with ExcelParser("data.xlsx") as parser:
            ...     tables = parser.parse_sheet("Sheet1")
        """
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """
        Context manager exit point.
        
        Automatically closes the parser when exiting the context manager,
        ensuring resources are properly released even if an exception occurs.
        
        Args:
            exc_type: Exception type (if any).
            exc_val: Exception value (if any).
            exc_tb: Exception traceback (if any).
        
        Examples:
            >>> with ExcelParser("data.xlsx") as parser:
            ...     tables = parser.parse_sheet("Sheet1")
            # Parser is automatically closed here
        """
        self.close()
    
    def __repr__(self):
        """
        String representation of the parser.
        
        Returns:
            str: A string representation showing the file path and parsing status.
        
        Examples:
            >>> parser = ExcelParser("data.xlsx")
            >>> print(parser)
            ExcelParser(file='data.xlsx', status='not parsed')
            >>> parser.parse_sheet("Sheet1")
            >>> print(parser)
            ExcelParser(file='data.xlsx', tables=3)
        """
        if self._tables is None:
            return f"ExcelParser(file='{self.file_path}', status='not parsed')"
        return f"ExcelParser(file='{self.file_path}', tables={len(self._tables)})"